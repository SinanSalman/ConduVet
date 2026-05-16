"""
Admin router — /api/admin/*
"""

import csv
import io
import os
import re
import shutil
from datetime import datetime, timedelta, timezone
from typing import Any

import yaml
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from passlib.context import CryptContext
from sqlalchemy import func
from sqlalchemy.orm import Session

from auth.jwt import create_access_token, get_current_admin
from database import get_db
from rate_limiter import limiter
from routers._helpers import log_field_changes as _log_field_changes, record_to_response as _record_to_response
from models.db_models import (
    AppConfig,
    AppUser,
    DataFile,
    DataRecord,
    FieldHistory,
    SchemaDefinition,
)
from services.excel_service import export_excel, parse_excel
from services.schema_parser import parse_data_type

router = APIRouter(prefix="/api/admin", tags=["admin"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

REQUIRED_YAML_KEYS = {"title", "admin_account", "admin_pass"}


def _validate_yaml(content: bytes) -> dict:
    try:
        data = yaml.safe_load(content)
    except yaml.YAMLError as exc:
        raise HTTPException(
            status_code=400,
            detail=(
                f"The configuration file contains invalid YAML syntax. "
                f"Common causes: incorrect indentation (use spaces, not tabs), "
                f"missing quotes around values that contain special characters (: [ ] {{ }}), "
                f"or a stray colon in a value. "
                f"Error detail: {exc}"
            ),
        )
    if not isinstance(data, dict):
        raise HTTPException(
            status_code=400,
            detail=(
                "The configuration file must be a YAML mapping (key: value pairs). "
                "It should look like this:\n"
                "  title: \"My Instance\"\n"
                "  admin_account: \"admin\"\n"
                "  admin_pass: \"secret\"\n"
                "  backup_dir: \"./backups\"\n"
                "Make sure the file does not start with a list (-) or a plain value."
            ),
        )
    missing = REQUIRED_YAML_KEYS - data.keys()
    if missing:
        pretty = ", ".join(f"'{k}'" for k in sorted(missing))
        raise HTTPException(
            status_code=400,
            detail=(
                f"The configuration file is missing required key(s): {pretty}. "
                f"Your config.yaml must include all of the following:\n"
                f"  title: \"Your instance title\"\n"
                f"  admin_account: \"admin_username\"\n"
                f"  admin_pass: \"admin_password\"\n"
                f"  backup_dir: \"./backups\"\n"
                f"Optional: auto_logout_minutes: 30 (default: 30, range: 1-480)"
            ),
        )

    # Validate admin_pass minimum length
    admin_pass = str(data.get("admin_pass", "")).strip()
    if len(admin_pass) < 6:
        raise HTTPException(
            status_code=400,
            detail=(
                "admin_pass must be at least 6 characters long. "
                "Choose a strong password to protect admin access."
            ),
        )

    # Validate optional auto_logout_minutes
    if "auto_logout_minutes" in data:
        logout_mins = data["auto_logout_minutes"]
        if not isinstance(logout_mins, int) or logout_mins < 1 or logout_mins > 480:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"auto_logout_minutes must be an integer between 1 and 480 (minutes). "
                    f"Got: {logout_mins}. This controls the session inactivity timeout."
                ),
            )
    else:
        # Set default if not specified
        data["auto_logout_minutes"] = 30

    # Validate optional user_domain
    if "user_domain" in data:
        user_domain = data["user_domain"]
        if not isinstance(user_domain, str) or not user_domain.strip():
            raise HTTPException(
                status_code=400,
                detail=(
                    f"user_domain must be a non-empty string. "
                    f"Example: user_domain: \"example.com\""
                ),
            )
        data["user_domain"] = user_domain.strip()
    else:
        data["user_domain"] = "example.com"

    # Validate optional pin_expiration_minutes
    if "pin_expiration_minutes" in data:
        expiry = data["pin_expiration_minutes"]
        if not isinstance(expiry, int) or expiry < 1 or expiry > 1440:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"pin_expiration_minutes must be an integer between 1 and 1440 (minutes). "
                    f"Got: {expiry}. This controls how long PINs remain valid."
                ),
            )
    else:
        data["pin_expiration_minutes"] = 15

    # Validate optional smtp_config
    if "smtp_config" in data:
        smtp = data["smtp_config"]
        if not isinstance(smtp, dict):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"smtp_config must be a YAML mapping with SMTP settings. "
                    f"Expected a dictionary, got: {type(smtp).__name__}. "
                    f"Example:\n"
                    f"  smtp_config:\n"
                    f"    host: \"smtp.gmail.com\"\n"
                    f"    port: 587\n"
                    f"    username: \"your-email@gmail.com\"\n"
                    f"    password: \"your-app-password\"\n"
                    f"    use_tls: true\n"
                    f"Note: password is optional for no-auth SMTP servers."
                ),
            )
        # Validate smtp_config required fields (host, port, username)
        # Password is optional for no-auth SMTP servers
        required_smtp_fields = {"host", "port", "username"}
        provided_fields = set(smtp.keys())
        if not required_smtp_fields.issubset(provided_fields):
            missing = required_smtp_fields - provided_fields
            pretty = ", ".join(f"'{k}'" for k in sorted(missing))
            raise HTTPException(
                status_code=400,
                detail=(
                    f"smtp_config is missing required field(s): {pretty}. "
                    f"All of the following must be specified:\n"
                    f"  host: \"smtp.gmail.com\"\n"
                    f"  port: 587\n"
                    f"  username: \"your-email@gmail.com\"\n"
                    f"Optional: password (for authenticated SMTP), use_tls (default: true)"
                ),
            )
        # Validate port is integer
        if not isinstance(smtp.get("port"), int):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"smtp_config.port must be an integer. "
                    f"Got: {smtp.get('port')}. Example: port: 587"
                ),
            )
        # Set defaults for optional fields
        if "use_tls" not in smtp:
            smtp["use_tls"] = True
        if "password" not in smtp:
            smtp["password"] = ""  # Empty password for no-auth SMTP
    else:
        # No SMTP config provided — set empty dict as default
        data["smtp_config"] = {}

    return data


def _parse_users_csv(content: bytes) -> list[dict]:
    """
    Parse a users CSV file.  Required columns: userid, name, password.
    Returns list of dicts with keys: userid (uppercased), name, password.
    """
    text = content.decode("utf-8-sig").strip()
    reader = csv.DictReader(io.StringIO(text))
    fieldnames_lower = [f.lower() for f in (reader.fieldnames or [])]
    required = {"userid", "name", "password"}
    missing = required - set(fieldnames_lower)
    if missing:
        pretty = ", ".join(f"'{c}'" for c in sorted(missing))
        found = ", ".join(f"'{f}'" for f in (reader.fieldnames or [])) or "(none detected)"
        raise HTTPException(
            status_code=400,
            detail=(
                f"The users CSV is missing required column(s): {pretty}. "
                f"Column(s) found in the file: {found}. "
                f"The first row of the CSV must be exactly:\n"
                f"  userid,name,password\n"
                f"Example rows:\n"
                f"  Z1234,Jane Smith,abc123\n"
                f"  Z5678,John Doe,xyz789\n"
                f"Note: column names are case-insensitive but must match exactly (no extra spaces)."
            ),
        )

    users = []
    skipped = 0
    for row in reader:
        row_lower = {k.lower(): v for k, v in row.items()}
        uid = (row_lower.get("userid") or "").strip()
        name = (row_lower.get("name") or "").strip()
        pwd = (row_lower.get("password") or "").strip()
        if uid:
            users.append({"userid": uid.upper(), "name": name, "password": pwd})
        else:
            skipped += 1

    if not users:
        raise HTTPException(
            status_code=400,
            detail=(
                "The users CSV contains no valid user rows. "
                "Check that the file has data rows beneath the header row, "
                "and that the 'userid' column is not empty."
            ),
        )
    return users


# ---------------------------------------------------------------------------
# GET /status — no auth
# ---------------------------------------------------------------------------

@router.get("/status")
def get_status(db: Session = Depends(get_db)):
    configured = db.query(AppConfig).first() is not None
    return {"configured": configured}


# ---------------------------------------------------------------------------
# POST /setup — no auth
# ---------------------------------------------------------------------------

@router.post("/setup")
async def setup(
    yaml_file: UploadFile = File(...),
    users_file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # Only allow setup once
    if db.query(AppConfig).first() is not None:
        raise HTTPException(status_code=400, detail="Application is already configured.")

    yaml_content = await yaml_file.read()
    config_data = _validate_yaml(yaml_content)

    users_content = await users_file.read()
    users = _parse_users_csv(users_content)

    # Save users file to disk
    users_dir = os.getenv("USERS_FILE_DIR", "/data/users")
    os.makedirs(users_dir, exist_ok=True)
    users_file_path = os.path.join(users_dir, "users.csv")
    with open(users_file_path, "wb") as fh:
        fh.write(users_content)

    backup_dir = os.getenv("BACKUP_DIR", "/data/backups")

    admin_pass_hash = pwd_context.hash(str(config_data["admin_pass"]))

    app_config = AppConfig(
        title=str(config_data["title"]),
        admin_account=str(config_data["admin_account"]),
        admin_pass_hash=admin_pass_hash,
        users_file_path=users_file_path,
        backup_dir=backup_dir,
        auto_logout_minutes=config_data.get("auto_logout_minutes", 30),
        user_domain=config_data.get("user_domain", "example.com"),
        pin_expiration_minutes=config_data.get("pin_expiration_minutes", 15),
        smtp_config=config_data.get("smtp_config", {}),
    )
    db.add(app_config)

    # Create/replace all app users
    db.query(AppUser).delete()
    for u in users:
        pwd_hash = pwd_context.hash(u["password"]) if u["password"] else pwd_context.hash("")
        db.add(AppUser(userid=u["userid"], name=u["name"], password_hash=pwd_hash))

    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# POST /login — no auth
# ---------------------------------------------------------------------------

@router.post("/login")
@limiter.limit("10/minute")
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    config = db.query(AppConfig).first()
    if config is None:
        raise HTTPException(status_code=503, detail="not_configured")

    if username != config.admin_account or not pwd_context.verify(
        password, config.admin_pass_hash
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid admin credentials",
        )

    token = create_access_token(
        data={"sub": username.upper(), "scope": "admin"},
        expires_delta=timedelta(hours=8),
    )
    return {
        "access_token": token,
        "token_type": "bearer",
        "title": config.title,
    }


# ---------------------------------------------------------------------------
# GET /config — admin auth
# ---------------------------------------------------------------------------

@router.get("/config")
def get_config(
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    config = db.query(AppConfig).first()
    if config is None:
        raise HTTPException(status_code=404, detail="Not configured")
    return {
        "id": config.id,
        "title": config.title,
        "admin_account": config.admin_account,
        "users_file_path": config.users_file_path,
        "backup_dir": config.backup_dir,
        "auto_logout_minutes": config.auto_logout_minutes,
        "user_domain": config.user_domain or "example.com",
        "pin_expiration_minutes": config.pin_expiration_minutes or 15,
        "smtp_config": config.smtp_config or {},
        "created_at": config.created_at.isoformat() if config.created_at else None,
        "updated_at": config.updated_at.isoformat() if config.updated_at else None,
    }


# ---------------------------------------------------------------------------
# POST /config/yaml — update configuration YAML only (admin auth)
# ---------------------------------------------------------------------------

@router.post("/config/yaml")
async def update_config_yaml(
    yaml_file: UploadFile = File(...),
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Update the application configuration (title, admin credentials) from a YAML file."""
    config = db.query(AppConfig).first()
    if config is None:
        raise HTTPException(status_code=404, detail="Not configured")

    yaml_content = await yaml_file.read()
    config_data = _validate_yaml(yaml_content)

    config.title = str(config_data["title"])
    config.admin_account = str(config_data["admin_account"])
    config.admin_pass_hash = pwd_context.hash(str(config_data["admin_pass"]))
    if "auto_logout_minutes" in config_data:
        config.auto_logout_minutes = config_data["auto_logout_minutes"]
    if "user_domain" in config_data:
        config.user_domain = config_data["user_domain"]
    if "pin_expiration_minutes" in config_data:
        config.pin_expiration_minutes = config_data["pin_expiration_minutes"]
    if "smtp_config" in config_data:
        config.smtp_config = config_data["smtp_config"]

    db.commit()

    # Reload to get updated timestamps
    db.refresh(config)
    return {
        "ok": True,
        "title": config.title,
        "admin_account": config.admin_account,
        "auto_logout_minutes": config.auto_logout_minutes,
        "user_domain": config.user_domain or "example.com",
        "pin_expiration_minutes": config.pin_expiration_minutes or 15,
        "smtp_config": config.smtp_config or {},
    }


# ---------------------------------------------------------------------------
# POST /config/users — update users CSV only (admin auth)
# ---------------------------------------------------------------------------

@router.post("/config/users")
async def update_config_users(
    users_file: UploadFile = File(...),
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Replace all application users from a CSV file."""
    config = db.query(AppConfig).first()
    if config is None:
        raise HTTPException(status_code=404, detail="Not configured")

    users_content = await users_file.read()
    users = _parse_users_csv(users_content)

    # Persist updated users file to disk
    users_file_path = config.users_file_path or "/data/users/users.csv"
    users_dir = os.path.dirname(users_file_path)
    os.makedirs(users_dir, exist_ok=True)
    with open(users_file_path, "wb") as fh:
        fh.write(users_content)

    # Replace all user records
    db.query(AppUser).delete()
    for u in users:
        pwd_hash = pwd_context.hash(u["password"]) if u["password"] else pwd_context.hash("")
        db.add(AppUser(userid=u["userid"], name=u["name"], password_hash=pwd_hash))

    db.commit()
    return {"ok": True, "user_count": len(users)}


# ---------------------------------------------------------------------------
# GET /files — admin auth
# ---------------------------------------------------------------------------

@router.get("/files")
def list_files(
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    files = db.query(DataFile).filter(DataFile.is_active.is_(True)).all()
    result = []
    for f in files:
        count = (
            db.query(func.count(DataRecord.id))
            .filter(DataRecord.file_id == f.id)
            .scalar()
        )
        result.append(
            {
                "id": f.id,
                "filename": f.filename,
                "display_name": f.display_name,
                "uploaded_at": f.uploaded_at.isoformat() if f.uploaded_at else None,
                "is_active": f.is_active,
                "record_count": count,
            }
        )
    return result


# ---------------------------------------------------------------------------
# POST /files/upload — admin auth
# ---------------------------------------------------------------------------

@router.post("/files/upload")
async def upload_file(
    file: UploadFile = File(...),
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    file_bytes = await file.read()
    filename = file.filename or "upload.xlsx"

    # Check if a file with this name already exists
    existing_file = db.query(DataFile).filter(
        DataFile.filename == filename,
        DataFile.is_active == True,
    ).first()
    if existing_file:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"A file named '{filename}' already exists on the server. "
                f"Please remove the existing file before uploading a new one with the same name. "
                f"Go to the Files tab, find '{existing_file.display_name}', and click Remove."
            ),
        )

    try:
        data_rows, schema_list, history_rows = parse_excel(file_bytes)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    # Derive display name from filename (strip extension)
    display_name = re.sub(r"\.[^.]+$", "", filename).strip() or filename

    data_file = DataFile(filename=filename, display_name=display_name)
    db.add(data_file)
    db.flush()  # get data_file.id

    # Persist schema definitions, excluding system columns.
    # Owner, Vetter, Last Updated, and Record Status are managed by dedicated DB
    # columns; storing them as SchemaDefinitions would cause duplicate grid columns.
    _SYSTEM_COLS = {"owner", "vetter", "record vetter", "last updated", "record status"}
    for s in schema_list:
        if s["field_name"].lower() in _SYSTEM_COLS:
            continue
        db.add(
            SchemaDefinition(
                file_id=data_file.id,
                field_name=s["field_name"],
                description=s["description"],
                data_type=s["data_type"],
                sample_data=s["sample_data"],
                depends_on=s["depends_on"],
                accept_null=s["accept_null"],
                is_protected=s.get("is_protected", False),
                field_order=s["field_order"],
            )
        )

    # Persist data records
    now = datetime.now(timezone.utc)
    _SYSTEM = {"owner", "vetter", "record vetter", "last updated", "record status", "record id"}
    schema_field_names = [
        s["field_name"] for s in schema_list
        if s["field_name"].lower() not in _SYSTEM
    ]

    records_list: list[DataRecord] = []   # references in the same order as data_rows
    old_ids_list: list[Any] = []          # "Record ID" from the Excel row (may be None)

    for row in data_rows:
        # Extract system columns (case-insensitive key match)
        row_lower = {k.lower(): v for k, v in row.items()}
        owner = str(row_lower.get("owner") or "ALL").strip().upper()
        # Vetter: stored uppercase like owner; check "record vetter" field; None if not present in the row
        vetter_raw = row_lower.get("record vetter") or row_lower.get("vetter")
        vetter = str(vetter_raw).strip().upper() if vetter_raw else None
        last_updated_raw = row_lower.get("last updated")
        record_status_raw = row_lower.get("record status")
        record_status = str(record_status_raw).strip() if record_status_raw else "Unvetted"

        # Parse last_updated
        last_updated = now
        if last_updated_raw:
            lu_str = str(last_updated_raw).strip()
            for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y"):
                try:
                    last_updated = datetime.strptime(lu_str, fmt).replace(tzinfo=timezone.utc)
                    break
                except ValueError:
                    pass

        # Build record_data from schema fields only, excluding system columns.
        # Owner, Vetter, Last Updated, Record Status are held in dedicated DB
        # columns and must not also live in the JSONB blob.
        record_data: dict[str, Any] = {}
        for field_name in schema_field_names:
            # Find matching column (case-insensitive)
            matched_val = None
            for col_name, col_val in row.items():
                if col_name.lower() == field_name.lower():
                    matched_val = col_val
                    break
            record_data[field_name] = matched_val

        record = DataRecord(
            file_id=data_file.id,
            owner=owner,
            vetter=vetter,
            record_data=record_data,
            record_status=record_status,
            last_updated=last_updated,
            created_at=now,
        )
        db.add(record)
        records_list.append(record)
        old_ids_list.append(row_lower.get("record id"))

    # Flush to assign IDs to all new records without committing the transaction.
    db.flush()

    # Build a mapping from the Excel "Record ID" values → newly assigned DB IDs.
    # This allows Edit History rows to be re-linked to the correct new records.
    old_id_to_new_id: dict[int, int] = {}
    for old_id_raw, record in zip(old_ids_list, records_list):
        if old_id_raw is not None:
            try:
                old_id_to_new_id[int(old_id_raw)] = record.id
            except (ValueError, TypeError):
                pass

    # Import Edit History rows from the workbook (if the sheet was present).
    for h in history_rows:
        old_rid = h.get("record_id")   # int or None
        if old_rid is not None:
            new_rid = old_id_to_new_id.get(old_rid)
            if new_rid is None:
                # History row references an ID not in the Data sheet.
                # For deletion events the record truly no longer exists — keep
                # the entry with record_id=NULL.  For other events, skip to
                # avoid orphaned history with no matching record.
                if h.get("field_name") != "_ROW_DELETED":
                    continue
                new_rid = None
        else:
            new_rid = None  # already NULL (e.g. previously deleted record)

        db.add(
            FieldHistory(
                record_id=new_rid,
                file_id=data_file.id,
                field_name=h["field_name"],
                old_value=h.get("old_value"),
                new_value=h.get("new_value"),
                changed_by=h.get("changed_by") or "imported",
                changed_at=h.get("changed_at") or now,
            )
        )

    record_count = len(records_list)
    db.commit()
    return {
        "id": data_file.id,
        "filename": data_file.filename,
        "display_name": data_file.display_name,
        "record_count": record_count,
    }


# ---------------------------------------------------------------------------
# DELETE /files/{file_id} — admin auth
# ---------------------------------------------------------------------------

@router.delete("/files/{file_id}")
def deactivate_file(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    data_file = db.query(DataFile).filter(DataFile.id == file_id).first()
    if data_file is None:
        raise HTTPException(status_code=404, detail="File not found")
    data_file.is_active = False
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# POST /reset — wipe all data and return to pre-setup state (admin auth)
# ---------------------------------------------------------------------------

@router.post("/reset")
def reset_all_data(
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """
    Delete every piece of application data and return the instance to its
    fresh state (as if the setup wizard had never been run).

    Deleted in dependency order so FK constraints are satisfied:
      1. FieldHistory          (references DataRecord via CASCADE — handled by ORM)
      2. DataRecord            (references DataFile via CASCADE — handled by ORM)
      3. SchemaDefinition      (references DataFile via CASCADE — handled by ORM)
      4. DataFile
      5. AppUser
      6. AppConfig
      7. On-disk: users CSV file (if stored path is still present)

    After this call the app will redirect all requests to the setup wizard.
    """
    try:
        # Fetch config before deleting it so we can clean up the users CSV path.
        config = db.query(AppConfig).order_by(AppConfig.id.desc()).first()
        users_file_path = config.users_file_path if config else None

        # Delete in strict dependency order so FK constraints are never violated.
        # Bulk deletes (synchronize_session=False) bypass ORM cascades, so each
        # child table must be cleared before its parent.
        db.query(FieldHistory).delete(synchronize_session=False)
        db.query(SchemaDefinition).delete(synchronize_session=False)
        db.query(DataRecord).delete(synchronize_session=False)
        db.query(DataFile).delete(synchronize_session=False)
        db.query(AppUser).delete(synchronize_session=False)
        db.query(AppConfig).delete(synchronize_session=False)

        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail="Reset failed — the database could not be cleared. Please try again.",
        )

    # Best-effort: remove the on-disk users CSV.  Not fatal if it is missing.
    if users_file_path:
        try:
            if os.path.isfile(users_file_path):
                os.remove(users_file_path)
        except OSError:
            pass

    return {"ok": True}


# ---------------------------------------------------------------------------
# GET /files/{file_id}/download — admin auth
# ---------------------------------------------------------------------------

@router.get("/files/{file_id}/download")
def download_file(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    data_file = db.query(DataFile).filter(DataFile.id == file_id).first()
    if data_file is None:
        raise HTTPException(status_code=404, detail="File not found")
    try:
        xlsx_bytes = export_excel(file_id, db)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    safe_name = re.sub(r"[^\w\-. ]", "_", data_file.display_name)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# GET /files/{file_id}/schema — admin auth
# ---------------------------------------------------------------------------

@router.get("/files/{file_id}/schema")
def get_schema(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    # Get the data file to include display_name in response
    data_file = db.query(DataFile).filter(DataFile.id == file_id).first()

    schemas = (
        db.query(SchemaDefinition)
        .filter(SchemaDefinition.file_id == file_id)
        .order_by(SchemaDefinition.field_order)
        .all()
    )
    return {
        "display_name": data_file.display_name if data_file else None,
        "fields": [
            {
                "id": s.id,
                "file_id": s.file_id,
                "field_name": s.field_name,
                "description": s.description,
                "data_type": s.data_type,
                "sample_data": s.sample_data,
                "depends_on": s.depends_on,
                "accept_null": s.accept_null,
                "is_protected": s.is_protected,
                "field_order": s.field_order,
            }
            for s in schemas
        ]
    }


# ---------------------------------------------------------------------------
# GET /files/{file_id}/records — admin auth (ALL records)
# ---------------------------------------------------------------------------

@router.get("/files/{file_id}/records")
def get_all_records(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    records = (
        db.query(DataRecord)
        .filter(DataRecord.file_id == file_id)
        .order_by(DataRecord.id)
        .all()
    )
    return [_record_to_response(r) for r in records]


# ---------------------------------------------------------------------------
# PUT /files/{file_id}/records — admin auth — batch update
# ---------------------------------------------------------------------------

@router.put("/files/{file_id}/records")
def batch_update_records(
    file_id: int,
    updates: list[dict],
    _admin: dict = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    changed_by = _admin.get("sub", "admin")
    now = datetime.now(timezone.utc)
    updated_ids = []

    for update in updates:
        record_id = update.get("id")
        if record_id is None:
            continue
        record = (
            db.query(DataRecord)
            .filter(DataRecord.id == record_id, DataRecord.file_id == file_id)
            .first()
        )
        if record is None:
            continue

        new_data = update.get("data", {})
        _log_field_changes(db, record, new_data, changed_by, now)

        record.record_data = new_data
        if "record_status" in update:
            old_status = record.record_status
            new_status = update["record_status"]
            if new_status != old_status:
                db.add(FieldHistory(
                    record_id=record.id,
                    file_id=record.file_id,
                    field_name="Record Status",
                    old_value=old_status,
                    new_value=new_status,
                    changed_by=changed_by,
                    changed_at=now,
                ))
            record.record_status = new_status
        if "owner" in update:
            record.owner = str(update["owner"]).upper()
        # Handle both "vetter" and "record vetter" field names
        vetter_update = update.get("record vetter") or update.get("vetter")
        if vetter_update is not None:
            record.vetter = str(vetter_update).strip().upper() if vetter_update else None
        record.last_updated = now
        updated_ids.append(record_id)

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to save records. Please try again.")
    return {"ok": True, "updated": len(updated_ids)}


# ---------------------------------------------------------------------------
# GET /files/{file_id}/records/{record_id}/history/{field_name} — admin auth
# ---------------------------------------------------------------------------

@router.get("/files/{file_id}/records/{record_id}/history/{field_name}")
def get_field_history_admin(
    file_id: int,
    record_id: int,
    field_name: str,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    history = (
        db.query(FieldHistory)
        .filter(
            FieldHistory.record_id == record_id,
            FieldHistory.file_id == file_id,
            FieldHistory.field_name == field_name,
        )
        .order_by(FieldHistory.changed_at.desc())
        .all()
    )

    # Resolve userids to display names in one query
    userids = list({h.changed_by for h in history if h.changed_by})
    name_map: dict[str, str] = {}
    if userids:
        rows = db.query(AppUser.userid, AppUser.name).filter(AppUser.userid.in_(userids)).all()
        name_map = {r.userid: r.name for r in rows}

    return [
        {
            "id": h.id,
            "record_id": h.record_id,
            "file_id": h.file_id,
            "field_name": h.field_name,
            "old_value": h.old_value,
            "new_value": h.new_value,
            "changed_by": h.changed_by,
            "changed_by_name": name_map.get(h.changed_by, h.changed_by),
            "changed_at": h.changed_at.isoformat() if h.changed_at else None,
        }
        for h in history
    ]


# ---------------------------------------------------------------------------
# Report helpers
# ---------------------------------------------------------------------------

def _get_by_user_data(file_id: int, db: Session) -> list[dict]:
    """Records grouped by userid (owner), showing updated/added counts."""
    records = (
        db.query(DataRecord)
        .filter(DataRecord.file_id == file_id)
        .all()
    )
    groups: dict[str, dict] = {}
    for r in records:
        uid = r.owner
        if uid not in groups:
            groups[uid] = {"owner": uid, "total": 0, "new": 0, "updated": 0,
                           "old": 0, "unvetted": 0, "delete": 0}
        groups[uid]["total"] += 1
        status_lower = (r.record_status or "").lower()
        if status_lower == "new":
            groups[uid]["new"] += 1
        elif status_lower == "updated":
            groups[uid]["updated"] += 1
        elif status_lower == "old":
            groups[uid]["old"] += 1
        elif status_lower == "unvetted":
            groups[uid]["unvetted"] += 1
        elif status_lower == "delete":
            groups[uid]["delete"] += 1
    return list(groups.values())


def _get_by_record_data(file_id: int, db: Session) -> list[dict]:
    """Records with their full field history summary."""
    records = (
        db.query(DataRecord)
        .filter(DataRecord.file_id == file_id)
        .order_by(DataRecord.id)
        .all()
    )
    result = []
    for r in records:
        history_count = (
            db.query(func.count(FieldHistory.id))
            .filter(FieldHistory.record_id == r.id)
            .scalar()
        )
        result.append(
            {
                "id": r.id,
                "owner": r.owner,
                "record_status": r.record_status,
                "last_updated": r.last_updated.strftime("%d-%m-%y %H:%M:%S") if r.last_updated else None,
                "field_changes": history_count,
            }
        )
    return result



def _build_report_xlsx(headers: list[str], rows: list[list]) -> bytes:
    """Build a simple single-sheet Excel report."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(h) + 4)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# GET /reports/{file_id}/by-user — admin auth
# ---------------------------------------------------------------------------

@router.get("/reports/{file_id}/by-user")
def report_by_user(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    data = _get_by_user_data(file_id, db)
    columns = ["Owner", "Total", "New", "Updated", "Old", "Unvetted", "Delete"]
    rows = [
        [d["owner"], d["total"], d["new"], d["updated"], d["old"], d["unvetted"], d["delete"]]
        for d in data
    ]
    return {"columns": columns, "rows": rows}


@router.get("/reports/{file_id}/by-user/download")
def report_by_user_download(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    data = _get_by_user_data(file_id, db)
    headers = ["Owner", "Total", "New", "Updated", "Old", "Delete"]
    rows = [
        [d["owner"], d["total"], d["new"], d["updated"], d["old"], d["unvetted"], d["delete"]]
        for d in data
    ]
    xlsx_bytes = _build_report_xlsx(headers, rows)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_by_user.xlsx"'},
    )


# ---------------------------------------------------------------------------
# GET /reports/{file_id}/by-record — admin auth
# ---------------------------------------------------------------------------

@router.get("/reports/{file_id}/by-record")
def report_by_record(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    data = _get_by_record_data(file_id, db)
    columns = ["ID", "Owner", "Record Status", "Last Updated", "Field Changes"]
    rows = [
        [d["id"], d["owner"], d["record_status"], d["last_updated"], d["field_changes"]]
        for d in data
    ]
    return {"columns": columns, "rows": rows}


@router.get("/reports/{file_id}/by-record/download")
def report_by_record_download(
    file_id: int,
    _admin=Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    data = _get_by_record_data(file_id, db)
    headers = ["ID", "Owner", "Record Status", "Last Updated", "Field Changes"]
    rows = [
        [d["id"], d["owner"], d["record_status"], d["last_updated"], d["field_changes"]]
        for d in data
    ]
    xlsx_bytes = _build_report_xlsx(headers, rows)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_by_record.xlsx"'},
    )


