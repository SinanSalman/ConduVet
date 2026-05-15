"""
Excel parsing and export using openpyxl only (no pandas).
"""

import io
from datetime import datetime, date, time
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Columns that are system-managed and do not need to appear in the Schema sheet
SYSTEM_COLUMNS = {"owner", "last updated", "record status"}

# Required Schema sheet column headers (case-insensitive matching)
SCHEMA_HEADERS = [
    "field name",
    "description",
    "data type",
    "sample data",
    "depends on",
    "accept null values",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_value(cell) -> Any:
    """Return a cleaned, JSON-serialisable cell value (strip strings, keep None for empty).

    openpyxl returns date/datetime/time cells as Python datetime objects.  These are
    not JSON-serialisable and would crash the JSONB insert.  We normalise them to
    DD/MM/YYYY, DD/MM/YYYY HH:MM:SS, and HH:MM:SS strings respectively.
    """
    val = cell.value
    if val is None:
        return None
    if isinstance(val, datetime):
        # datetime must come before date because datetime is a subclass of date
        return val.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def _to_bool(val) -> bool:
    if val is None:
        return True
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ("yes", "true", "1", "y")


def _date_to_display(value: Any) -> Optional[str]:
    """Convert ISO 8601 date/datetime string or date object to DD/MM/YYYY."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    s = str(value).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return s  # return as-is if we cannot parse


def _datetime_to_display(value: Any) -> Optional[str]:
    """Convert ISO 8601 datetime string or object to DD/MM/YYYY HH:MM:SS."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    s = str(value).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y %H:%M:%S")
        except ValueError:
            pass
    return s


# ---------------------------------------------------------------------------
# parse_excel
# ---------------------------------------------------------------------------

def _parse_history_datetime(raw) -> Optional[datetime]:
    """Try to parse a cell value as a datetime for the Edit History sheet."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day)
    s = str(raw).strip()
    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Validate and parse an uploaded Excel file.

    Returns:
        (data_rows, schema_list, history_rows)

        data_rows   : list of dicts — one dict per data row, keys = column headers
        schema_list : list of dicts — one dict per schema field with keys:
                         field_name, description, data_type, sample_data,
                         depends_on, accept_null, field_order
        history_rows: list of dicts — one dict per history entry from the optional
                         "Edit History" sheet.  Empty list if the sheet is absent.
                         Keys: record_id (int|None), field_name, old_value,
                               new_value, changed_by, changed_at (datetime|None)

    Raises:
        ValueError with a descriptive message on invalid format.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(
            f"The file could not be opened as a valid Excel workbook. "
            f"Make sure it is saved in .xlsx format (not .xls, .csv, or a renamed file), "
            f"is not password-protected, and is not corrupted. "
            f"Technical detail: {exc}"
        ) from exc

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
    found_sheets = ", ".join(f"'{s}'" for s in wb.sheetnames) if wb.sheetnames else "(none)"

    if "data" not in sheet_names_lower:
        raise ValueError(
            f"This workbook is missing the required 'Data' sheet. "
            f"Sheet(s) found: {found_sheets}. "
            f"Please rename your data sheet to exactly 'Data' and re-upload."
        )
    if "schema" not in sheet_names_lower:
        raise ValueError(
            f"This workbook is missing the required 'Schema' sheet. "
            f"Sheet(s) found: {found_sheets}. "
            f"Please add a 'Schema' sheet with field definitions (columns: "
            f"Field Name, Description, Data Type, Sample Data, Depends on, Accept Null Values) "
            f"and re-upload."
        )

    ws_data = wb[sheet_names_lower["data"]]
    ws_schema = wb[sheet_names_lower["schema"]]

    # ---- Parse Schema sheet ------------------------------------------------
    schema_rows = list(ws_schema.iter_rows(values_only=False))
    if not schema_rows:
        raise ValueError("Schema sheet is empty.")

    # Find header row (first non-empty row)
    schema_header_map: dict[str, int] = {}
    header_row_idx = None
    for row_idx, row in enumerate(schema_rows):
        non_empty = [c for c in row if _cell_value(c) is not None]
        if non_empty:
            header_row_idx = row_idx
            for col_idx, cell in enumerate(row):
                val = _cell_value(cell)
                if val is not None:
                    schema_header_map[str(val).lower()] = col_idx
            break

    if header_row_idx is None:
        raise ValueError(
            "The 'Schema' sheet appears to be empty. It must have a header row as its "
            "first non-empty row with these columns (in any order): "
            "Field Name, Description, Data Type, Sample Data, Depends on, Accept Null Values."
        )

    # Verify all required headers are present
    missing_headers = [h for h in SCHEMA_HEADERS if h not in schema_header_map]
    if missing_headers:
        pretty_missing = ", ".join(f"'{h.title()}'" for h in missing_headers)
        pretty_expected = ", ".join(f"'{h.title()}'" for h in SCHEMA_HEADERS)
        raise ValueError(
            f"The 'Schema' sheet header row is missing required column(s): {pretty_missing}. "
            f"All of the following columns must be present (spelling and spacing must match exactly): "
            f"{pretty_expected}. "
            f"Check for extra spaces, merged cells, or typos in the header row."
        )

    schema_list: list[dict] = []
    for row_idx in range(header_row_idx + 1, len(schema_rows)):
        row = schema_rows[row_idx]
        field_name = _cell_value(row[schema_header_map["field name"]])
        if field_name is None:
            continue  # skip blank rows

        description = _cell_value(row[schema_header_map["description"]])
        data_type = _cell_value(row[schema_header_map["data type"]]) or "Text"
        sample_data = _cell_value(row[schema_header_map["sample data"]])
        depends_on = _cell_value(row[schema_header_map["depends on"]])
        accept_null_raw = _cell_value(row[schema_header_map["accept null values"]])
        accept_null = _to_bool(accept_null_raw)

        # Parse optional "Protected" column (for field-level protection)
        is_protected_raw = None
        if "protected" in schema_header_map:
            is_protected_raw = _cell_value(row[schema_header_map["protected"]])
        is_protected = _to_bool(is_protected_raw)

        schema_list.append(
            {
                "field_name": str(field_name),
                "description": str(description) if description is not None else "",
                "data_type": str(data_type),
                "sample_data": str(sample_data) if sample_data is not None else "",
                "depends_on": str(depends_on) if depends_on is not None else "",
                "accept_null": accept_null,
                "is_protected": is_protected,
                "field_order": len(schema_list),
            }
        )

    if not schema_list:
        raise ValueError(
            "The 'Schema' sheet has a header row but no field definition rows below it. "
            "Add at least one row beneath the header, with each row defining one data field."
        )

    # ---- Parse Data sheet --------------------------------------------------
    data_rows_raw = list(ws_data.iter_rows(values_only=False))
    if not data_rows_raw:
        raise ValueError(
            "The 'Data' sheet is empty. It must contain at least a header row listing "
            "the column names (including Owner, Last Updated, and Record Status)."
        )

    # Find header row
    data_header_map: dict[str, int] = {}
    data_header_row_idx = None
    for row_idx, row in enumerate(data_rows_raw):
        non_empty = [c for c in row if _cell_value(c) is not None]
        if non_empty:
            data_header_row_idx = row_idx
            for col_idx, cell in enumerate(row):
                val = _cell_value(cell)
                if val is not None:
                    data_header_map[str(val)] = col_idx
            break

    if data_header_row_idx is None:
        raise ValueError(
            "The 'Data' sheet appears to contain only empty rows. "
            "The first non-empty row must be a header row listing column names."
        )

    # Every schema field name must exist as a Data column (system columns exempt)
    schema_field_names = {s["field_name"] for s in schema_list}
    data_col_names_lower = {k.lower() for k in data_header_map}
    missing_cols = sorted(
        f
        for f in schema_field_names
        if f.lower() not in data_col_names_lower
        and f.lower() not in SYSTEM_COLUMNS
    )
    if missing_cols:
        quoted = ", ".join(f"'{c}'" for c in missing_cols)
        raise ValueError(
            f"The following field(s) are defined in the 'Schema' sheet but have no matching "
            f"column in the 'Data' sheet: {quoted}. "
            f"Either add these columns to the 'Data' sheet header row, or remove them from "
            f"the 'Schema' sheet. Column names must match exactly (including spaces and capitalisation)."
        )

    # Build column name lookup (case-insensitive)
    col_name_by_idx = {v: k for k, v in data_header_map.items()}

    data_rows: list[dict] = []
    for row_idx in range(data_header_row_idx + 1, len(data_rows_raw)):
        row = data_rows_raw[row_idx]
        # Skip fully empty rows
        if all(_cell_value(c) is None for c in row):
            continue
        row_dict: dict[str, Any] = {}
        for col_idx, col_name in col_name_by_idx.items():
            if col_idx < len(row):
                row_dict[col_name] = _cell_value(row[col_idx])
            else:
                row_dict[col_name] = None
        data_rows.append(row_dict)

    # ---- Parse Edit History sheet (optional) ---------------------------------
    history_rows: list[dict] = []
    if "edit history" in sheet_names_lower:
        ws_hist = wb[sheet_names_lower["edit history"]]
        hist_rows_raw = list(ws_hist.iter_rows(values_only=False))

        # Locate header row
        hist_header_map: dict[str, int] = {}
        hist_header_idx = None
        for row_idx, row in enumerate(hist_rows_raw):
            non_empty = [c for c in row if _cell_value(c) is not None]
            if non_empty:
                hist_header_idx = row_idx
                for col_idx, cell in enumerate(row):
                    val = _cell_value(cell)
                    if val is not None:
                        hist_header_map[str(val).lower().strip()] = col_idx
                break

        # Required column names (case-insensitive)
        _H_RECORD_ID  = "record id"
        _H_FIELD_NAME = "field name"
        _H_OLD_VALUE  = "old value"
        _H_NEW_VALUE  = "new value"
        _H_CHANGED_BY = "changed by"
        _H_CHANGED_AT = "changed at"

        if hist_header_idx is not None and all(
            h in hist_header_map
            for h in (_H_FIELD_NAME, _H_OLD_VALUE, _H_NEW_VALUE, _H_CHANGED_BY, _H_CHANGED_AT)
        ):
            for row_idx in range(hist_header_idx + 1, len(hist_rows_raw)):
                row = hist_rows_raw[row_idx]
                if all(_cell_value(c) is None for c in row):
                    continue

                def _hcell(key: str):
                    col = hist_header_map.get(key)
                    if col is None or col >= len(row):
                        return None
                    return _cell_value(row[col])

                # record_id may be NULL for deleted-record events
                raw_rid = _hcell(_H_RECORD_ID)
                try:
                    record_id = int(raw_rid) if raw_rid is not None else None
                except (ValueError, TypeError):
                    record_id = None

                field_name = _hcell(_H_FIELD_NAME)
                if not field_name:
                    continue  # skip rows without a field name

                history_rows.append({
                    "record_id":  record_id,
                    "field_name": str(field_name),
                    "old_value":  str(_hcell(_H_OLD_VALUE)) if _hcell(_H_OLD_VALUE) is not None else None,
                    "new_value":  str(_hcell(_H_NEW_VALUE)) if _hcell(_H_NEW_VALUE) is not None else None,
                    "changed_by": str(_hcell(_H_CHANGED_BY)) if _hcell(_H_CHANGED_BY) is not None else "imported",
                    "changed_at": _parse_history_datetime(_hcell(_H_CHANGED_AT)),
                })

    return data_rows, schema_list, history_rows


# ---------------------------------------------------------------------------
# export_excel
# ---------------------------------------------------------------------------

def export_excel(file_id: int, db) -> bytes:
    """
    Export a data file back to Excel bytes with Data + Schema + Edit History sheets.

    Dates stored as ISO 8601 internally are exported in DD/MM/YYYY format.
    """
    from models.db_models import DataFile, SchemaDefinition, DataRecord, FieldHistory
    from services.schema_parser import parse_data_type

    data_file = db.query(DataFile).filter(DataFile.id == file_id).first()
    if data_file is None:
        raise ValueError(f"No data file found with id={file_id}")

    schemas = (
        db.query(SchemaDefinition)
        .filter(SchemaDefinition.file_id == file_id)
        .order_by(SchemaDefinition.field_order)
        .all()
    )

    records = (
        db.query(DataRecord)
        .filter(DataRecord.file_id == file_id)
        .order_by(DataRecord.id)
        .all()
    )

    history_records = (
        db.query(FieldHistory)
        .filter(FieldHistory.file_id == file_id)
        .order_by(FieldHistory.changed_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()

    # ---- Data sheet --------------------------------------------------------
    ws_data = wb.active
    ws_data.title = "Data"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    # Build ordered data column list
    data_columns = [s.field_name for s in schemas] + [
        "Owner",
        "Last Updated",
        "Record Status",
        "Record ID",
    ]

    # Write header row
    for col_idx, col_name in enumerate(data_columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max(
            15, len(col_name) + 4
        )

    # Determine date/datetime fields for display formatting
    type_by_field: dict[str, str] = {}
    for s in schemas:
        parsed = parse_data_type(s.data_type)
        type_by_field[s.field_name] = parsed["type"]

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        rdata = record.record_data or {}
        col_idx = 1
        for field_name in [s.field_name for s in schemas]:
            raw_val = rdata.get(field_name)
            dtype = type_by_field.get(field_name, "text")
            if dtype == "date" and raw_val is not None:
                display_val = _date_to_display(raw_val)
            elif dtype == "datetime" and raw_val is not None:
                display_val = _datetime_to_display(raw_val)
            else:
                display_val = raw_val
            ws_data.cell(row=row_idx, column=col_idx, value=display_val)
            col_idx += 1

        # System columns
        ws_data.cell(row=row_idx, column=col_idx, value=record.owner)
        col_idx += 1
        lu = record.last_updated
        ws_data.cell(
            row=row_idx,
            column=col_idx,
            value=lu.strftime("%d/%m/%Y %H:%M:%S") if lu else None,
        )
        col_idx += 1
        ws_data.cell(row=row_idx, column=col_idx, value=record.record_status)
        col_idx += 1
        ws_data.cell(row=row_idx, column=col_idx, value=record.id)

    # ---- Schema sheet ------------------------------------------------------
    ws_schema = wb.create_sheet(title="Schema")

    schema_headers = [
        "Field Name",
        "Description",
        "Data Type",
        "Sample Data",
        "Depends on",
        "Accept Null Values",
    ]
    for col_idx, col_name in enumerate(schema_headers, start=1):
        cell = ws_schema.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws_schema.column_dimensions[get_column_letter(col_idx)].width = max(
            18, len(col_name) + 4
        )

    for row_idx, schema in enumerate(schemas, start=2):
        ws_schema.cell(row=row_idx, column=1, value=schema.field_name)
        ws_schema.cell(row=row_idx, column=2, value=schema.description)
        ws_schema.cell(row=row_idx, column=3, value=schema.data_type)
        ws_schema.cell(row=row_idx, column=4, value=schema.sample_data)
        ws_schema.cell(row=row_idx, column=5, value=schema.depends_on)
        ws_schema.cell(
            row=row_idx, column=6, value="Yes" if schema.accept_null else "No"
        )

    # ---- Edit History sheet ------------------------------------------------
    ws_history = wb.create_sheet(title="Edit History")

    history_headers = [
        "Record ID",
        "Field Name",
        "Old Value",
        "New Value",
        "Changed By",
        "Changed At",
    ]
    for col_idx, col_name in enumerate(history_headers, start=1):
        cell = ws_history.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws_history.column_dimensions[get_column_letter(col_idx)].width = max(
            18, len(col_name) + 4
        )

    for row_idx, history in enumerate(history_records, start=2):
        # For _ROW_DELETED events the record has been removed, so record_id is NULL.
        # The original record ID was stored in old_value at deletion time.
        display_record_id = history.record_id
        if display_record_id is None and history.field_name == "_ROW_DELETED":
            # Recover the original record ID from old_value if possible
            try:
                display_record_id = int(history.old_value) if history.old_value else None
            except (ValueError, TypeError):
                display_record_id = None

        ws_history.cell(row=row_idx, column=1, value=display_record_id)
        ws_history.cell(row=row_idx, column=2, value=history.field_name)
        ws_history.cell(row=row_idx, column=3, value=history.old_value)
        ws_history.cell(row=row_idx, column=4, value=history.new_value)
        ws_history.cell(row=row_idx, column=5, value=history.changed_by)
        changed_at = history.changed_at
        ws_history.cell(
            row=row_idx,
            column=6,
            value=changed_at.strftime("%d/%m/%Y %H:%M:%S") if changed_at else None,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
